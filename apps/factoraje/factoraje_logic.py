
"""
Translation of app.R (Shiny) to Python.
Goal: preserve the same logic / calculations, only translating language + UI to Streamlit.

Implements:
- list_invoices_headers_api (API headers)
- syntage_list_invoices_ids + XML download pipeline
- metrics_by_interval
- Excel export with grouped headers (similar layout)
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any, Dict, Iterable, List, Optional, Tuple

import json
import re
import threading

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =======================
# Helpers (mirrors R)
# =======================

def is_rfc(x: str) -> bool:
    if not isinstance(x, str):
        return False
    x = x.strip().upper()
    return bool(x) and bool(re.fullmatch(r"[A-Z0-9]{12,13}", x))

def today_utc() -> date:
    # R uses Sys.Date(); treat as local date.
    return date.today()

INTERVAL_DEFS: Dict[str, int] = {
    "Últimos 12 meses": 365,
    "Últimos 6 meses": 183,
    "Últimos 3 meses": 92,
    "Último mes": 30,
}

# Very light global monitor of last HTTP call
_api_mon_lock = threading.Lock()
_api_mon: Dict[str, Any] = {"last_status": None, "last_url": None, "last_ok": False}

def _set_api_mon(ok: bool, status: Optional[int], url: Optional[str]) -> None:
    with _api_mon_lock:
        _api_mon["last_ok"] = ok
        _api_mon["last_status"] = status
        _api_mon["last_url"] = url

def get_api_mon() -> Dict[str, Any]:
    with _api_mon_lock:
        return dict(_api_mon)


def _make_session() -> requests.Session:
    """
    Wraps basic retry behavior similar to httr2::req_retry + backoff.
    """
    sess = requests.Session()
    retry = Retry(
        total=5,
        backoff_factor=1.0,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["GET", "HEAD"],
        raise_on_status=False,
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20)
    sess.mount("http://", adapter)
    sess.mount("https://", adapter)
    return sess


def is_xml_ct(content_type: Optional[str]) -> bool:
    if not content_type:
        return False
    ct = content_type.lower()
    return ("xml" in ct) or (ct.startswith("text/") and "xml" in ct)


def _safe_get(d: Any, *keys: str) -> Optional[Any]:
    """
    Mirrors the R helper get(...) for a list/dict payload.
    Tries multiple keys and returns the first non-empty.
    """
    if not isinstance(d, dict):
        return None
    for k in keys:
        if k in d and d[k] is not None:
            v = d[k]
            # hydra often returns scalars; sometimes lists
            if isinstance(v, list) and v:
                return v[0]
            return v
    return None

def _safe_get_nested(d: Any, path: List[str]) -> Optional[Any]:
    cur = d
    for k in path:
        if not isinstance(cur, dict):
            return None
        cur = cur.get(k)
        if cur is None:
            return None
    if isinstance(cur, list) and cur:
        return cur[0]
    return cur


# =========================================================
# API: listar facturas (recibidas) y normalizar headers
# =========================================================

def list_invoices_headers_api(
    base_url: str,
    api_key: str,
    rfc: str,
    date_from: date,
    date_to: date,
    items_per_page: int = 1000,
    max_pages: int = 400,
    session: Optional[requests.Session] = None,
) -> pd.DataFrame:
    base = re.sub(r"/+$", "", base_url)
    url0 = f"{base}/taxpayers/{rfc}/invoices"

    qs_base = {
        "itemsPerPage": min(items_per_page, 1000),
        "isIssuer": "false",  # SOLO recibidas
        "issuedAt[after]": f"{date_from:%Y-%m-%d}T00:00:00Z",
        "issuedAt[before]": f"{date_to:%Y-%m-%d}T23:59:59Z",
    }

    sess = session or _make_session()
    headers = {"X-API-Key": api_key}

    acc: List[Dict[str, Any]] = []
    next_id_lt: Optional[str] = None

    for _ in range(max_pages):
        qs = dict(qs_base)
        if next_id_lt:
            qs["id[lt]"] = next_id_lt

        try:
            resp = sess.get(url0, headers=headers, params=qs, timeout=35)
            _set_api_mon(True, resp.status_code, resp.url)
        except Exception:
            _set_api_mon(False, None, url0)
            break

        if not (200 <= resp.status_code < 300):
            break

        try:
            j = resp.json()
        except Exception:
            break

        rows = j.get("hydra:member")
        if not rows:
            break
        if isinstance(rows, dict):
            rows = [rows]
        if isinstance(rows, list):
            acc.extend(rows)

        # pagination token: last id-ish field
        last_ids: List[str] = []
        for x in rows:
            if not isinstance(x, dict):
                continue
            for f in ("id", "@id", "invoiceId", "uuid", "documentId"):
                v = x.get(f)
                if isinstance(v, list):
                    v = v[0] if v else None
                if v is not None and str(v).strip():
                    last_ids.append(str(v))
                    break

        if not last_ids:
            break
        next_id_lt = last_ids[-1]

        if len(rows) < qs_base["itemsPerPage"]:
            break

    if not acc:
        return pd.DataFrame()

    # ---- Normalizar campos clave (same keys & fallbacks as R)
    norm_rows: List[Dict[str, Any]] = []

    for it in acc:
        if not isinstance(it, dict):
            continue

        uuid = _safe_get(it, "uuid", "id", "@id", "invoiceId", "documentId")
        issued = _safe_get(it, "issuedAt", "issueDate", "date", "createdAt")
        fecha = None
        if issued:
            try:
                fecha = datetime.fromisoformat(str(issued)[:10]).date()
            except Exception:
                fecha = None

        def _num(x: Any) -> Optional[float]:
            try:
                if x is None:
                    return None
                return float(str(x).replace(",", ""))
            except Exception:
                return None

        total = _num(_safe_get(it, "total", "totalAmount", "amount_total", "grandTotal", "importe_total"))
        moneda = _safe_get(it, "currency", "moneda")
        tipo_cambio = _num(_safe_get(it, "exchangeRate", "tipoCambio", "tipo_cambio"))
        metodo = _safe_get(it, "paymentMethod", "paymentType", "metodoPago")
        tipo = _safe_get(it, "type", "tipoDeComprobante", "tipo_de_comprobante")

        emisor_rfc = _safe_get_nested(it, ["issuer", "rfc"]) or _safe_get(it, "issuer.rfc", "emitter.rfc", "issuer_tax_id")
        emisor_nombre = _safe_get_nested(it, ["issuer", "name"]) or _safe_get(it, "issuer.name", "emitter.name")
        receptor_rfc = _safe_get_nested(it, ["receiver", "rfc"]) or _safe_get(it, "receiver.rfc", "receiver_tax_id")
        receptor_nombre = _safe_get_nested(it, ["receiver", "name"]) or _safe_get(it, "receiver.name")

        norm_rows.append(
            {
                "uuid": (str(uuid) if uuid is not None else None),
                "fecha": fecha,
                "total": total,
                "moneda": (str(moneda).upper() if moneda is not None else None),
                "tipo_cambio": tipo_cambio,
                "metodo": (re.sub(r"\s+", "", str(metodo)).upper() if metodo is not None else None),
                "tipo": (re.sub(r"\s+", "", str(tipo)).upper() if tipo is not None else None),
                "emisor_rfc": (str(emisor_rfc).upper() if emisor_rfc is not None else None),
                "emisor_nombre": (str(emisor_nombre) if emisor_nombre is not None else None),
                "receptor_rfc": (str(receptor_rfc).upper() if receptor_rfc is not None else None),
                "receptor_nombre": (str(receptor_nombre) if receptor_nombre is not None else None),
            }
        )

    h = pd.DataFrame(norm_rows)

    # ---- Filtros ESTRICTOS p/ “proveedores”
    if h.empty:
        return h

    rfc_u = rfc.upper()
    h = h[
        (h["uuid"].notna()) & (h["uuid"].astype(str) != "")
        & (h["receptor_rfc"].fillna("").str.upper() == rfc_u)
        & (h["emisor_rfc"].fillna("").str.upper() != rfc_u)
        & (h["tipo"].fillna("").str.upper() == "I")
    ].drop_duplicates(subset=["uuid"], keep="first")

    return h.reset_index(drop=True)


# =========================================================
# XML: listar IDs, descargar XML y parsear header mínimo
# =========================================================

def syntage_list_invoices_ids(
    base_url: str,
    api_key: str,
    rfc: str,
    date_from: date,
    date_to: date,
    items_per_page: int = 1000,
    max_pages: int = 400,
    session: Optional[requests.Session] = None,
) -> List[Dict[str, Any]]:
    # Mirrors R; returns raw list items
    base = re.sub(r"/+$", "", base_url)
    url0 = f"{base}/taxpayers/{rfc}/invoices"
    qs_base = {
        "itemsPerPage": min(items_per_page, 1000),
        "isIssuer": "false",
        "issuedAt[after]": f"{date_from:%Y-%m-%d}T00:00:00Z",
        "issuedAt[before]": f"{date_to:%Y-%m-%d}T23:59:59Z",
    }

    sess = session or _make_session()
    headers = {"X-API-Key": api_key}

    acc: List[Dict[str, Any]] = []
    next_id_lt: Optional[str] = None

    for _ in range(max_pages):
        qs = dict(qs_base)
        if next_id_lt:
            qs["id[lt]"] = next_id_lt

        try:
            resp = sess.get(url0, headers=headers, params=qs, timeout=35)
            _set_api_mon(True, resp.status_code, resp.url)
        except Exception:
            _set_api_mon(False, None, url0)
            break

        if not (200 <= resp.status_code < 300):
            break

        try:
            j = resp.json()
        except Exception:
            break

        rows = j.get("hydra:member")
        if not rows:
            break
        if isinstance(rows, dict):
            rows = [rows]
        if isinstance(rows, list):
            for x in rows:
                if isinstance(x, dict):
                    acc.append(x)

        last_ids: List[str] = []
        for x in rows:
            if not isinstance(x, dict):
                continue
            for f in ("id", "@id", "invoiceId", "uuid", "documentId"):
                v = x.get(f)
                if isinstance(v, list):
                    v = v[0] if v else None
                if v is not None and str(v).strip():
                    last_ids.append(str(v))
                    break
        if not last_ids:
            break
        next_id_lt = last_ids[-1]

        if len(rows) < qs_base["itemsPerPage"]:
            break

    return acc


def cfdi_url_candidates(base_url: str, it: Dict[str, Any]) -> List[str]:
    base = re.sub(r"/+$", "", base_url)
    cands: List[str] = []

    id_at = it.get("@id")
    if isinstance(id_at, list):
        id_at = id_at[0] if id_at else None
    if id_at:
        p = str(id_at)
        if p:
            if not p.startswith("/"):
                p = "/" + p
            cands.append(f"{base}{p}/cfdi")

    for f in ("uuid", "invoiceId", "documentId", "id"):
        v = it.get(f)
        if isinstance(v, list):
            v = v[0] if v else None
        if v:
            vv = str(v)
            if vv:
                cands.append(f"{base}/invoices/{vv}/cfdi")

    # unique, keep order
    seen = set()
    out: List[str] = []
    for u in cands:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out


def probe_first_working_url(
    urls: Iterable[str],
    api_key: str,
    timeout_secs: int = 8,
    session: Optional[requests.Session] = None,
) -> Optional[str]:
    sess = session or _make_session()
    headers = {"X-API-Key": api_key}
    for u in urls:
        try:
            resp = sess.head(u, headers=headers, timeout=timeout_secs, allow_redirects=True)
            _set_api_mon(True, resp.status_code, resp.url)
            if 200 <= resp.status_code < 300:
                return u
        except Exception:
            _set_api_mon(False, None, u)
            continue
    return None


def http_get_xml_flex(
    url: str,
    api_key: str,
    timeout_secs: int = 15,
    session: Optional[requests.Session] = None,
) -> Optional[str]:
    sess = session or _make_session()
    headers = {
        "X-API-Key": api_key,
        "Accept": "application/xml, text/xml;q=0.9, */*;q=0.5",
        "User-Agent": "HayCash-Concentracion/1.0",
    }
    try:
        resp = sess.get(url, headers=headers, timeout=timeout_secs)
        _set_api_mon(True, resp.status_code, resp.url)
    except Exception:
        _set_api_mon(False, None, url)
        return None

    ct = resp.headers.get("content-type")
    if 200 <= resp.status_code < 300:
        body = resp.text
        if is_xml_ct(ct) or re.match(r"^\s*<\?xml", body or ""):
            return body
    return None


def download_xmls_parallel(
    urls: List[str],
    api_key: str,
    timeout_secs: int = 15,
    parallel: bool = True,
) -> List[str]:
    # R optionally uses future.apply; we use threads (I/O bound).
    sess = _make_session()

    def _one(u: str) -> Optional[str]:
        return http_get_xml_flex(u, api_key, timeout_secs=timeout_secs, session=sess)

    if not parallel or len(urls) <= 1:
        out = [_one(u) for u in urls]
        return [x for x in out if x]

    from concurrent.futures import ThreadPoolExecutor, as_completed

    out: List[str] = []
    with ThreadPoolExecutor(max_workers=min(16, max(4, len(urls)))) as ex:
        futs = {ex.submit(_one, u): u for u in urls}
        for f in as_completed(futs):
            try:
                v = f.result()
                if v:
                    out.append(v)
            except Exception:
                continue
    return out


def parse_header_min(xml_raw: str) -> pd.DataFrame:
    try:
        doc = etree.fromstring(xml_raw.encode("utf-8", errors="ignore"))
    except Exception:
        return pd.DataFrame()

    def a1(xpath: str, attr: str) -> Optional[str]:
        try:
            node = doc.xpath(xpath)
            if not node:
                return None
            # xpath returns list of elements
            n0 = node[0]
            if isinstance(n0, etree._Element):
                return n0.get(attr)
            return None
        except Exception:
            return None

    def up(x: Optional[str]) -> str:
        return re.sub(r"\s+", "", (x or "")).upper()

    uuid = a1("//*[local-name()='TimbreFiscalDigital']", "UUID")
    tipo = up(a1("//*[local-name()='Comprobante']", "TipoDeComprobante"))
    if tipo not in {"I", "E", "P", "N", "T"}:
        tipo = None

    metodo = up(a1("//*[local-name()='Comprobante']", "MetodoPago"))
    fecha_s = a1("//*[local-name()='Comprobante']", "Fecha") or ""
    fecha = None
    try:
        fecha = datetime.fromisoformat(fecha_s[:10]).date()
    except Exception:
        fecha = None

    def _num_attr(v: Optional[str]) -> Optional[float]:
        try:
            if v is None:
                return None
            return float(str(v).replace(",", ""))
        except Exception:
            return None

    total = _num_attr(a1("//*[local-name()='Comprobante']", "Total"))
    moneda = up(a1("//*[local-name()='Comprobante']", "Moneda"))
    tipo_cambio = _num_attr(a1("//*[local-name()='Comprobante']", "TipoCambio"))
    emisor_rfc = up(a1("//*[local-name()='Emisor']", "Rfc"))
    emisor_nombre = a1("//*[local-name()='Emisor']", "Nombre")
    receptor_rfc = up(a1("//*[local-name()='Receptor']", "Rfc"))
    receptor_nombre = a1("//*[local-name()='Receptor']", "Nombre")

    return pd.DataFrame(
        [
            {
                "uuid": uuid,
                "tipo": tipo,
                "metodo": metodo,
                "fecha": fecha,
                "total": total,
                "moneda": moneda,
                "tipo_cambio": tipo_cambio,
                "emisor_rfc": emisor_rfc,
                "emisor_nombre": emisor_nombre,
                "receptor_rfc": receptor_rfc,
                "receptor_nombre": receptor_nombre,
            }
        ]
    )


def headers_from_xml(base_url: str, api_key: str, rfc: str, dfrom: date, dto: date) -> pd.DataFrame:
    items = syntage_list_invoices_ids(base_url, api_key, rfc, dfrom, dto)
    if not items:
        return pd.DataFrame()

    sess = _make_session()
    urls: List[str] = []
    for it in items:
        cands = cfdi_url_candidates(base_url, it)
        u = probe_first_working_url(cands, api_key, timeout_secs=8, session=sess)
        if u:
            urls.append(u)

    urls = [u for u in urls if u]
    if not urls:
        return pd.DataFrame()

    xmls = download_xmls_parallel(urls, api_key, parallel=True, timeout_secs=15)
    xmls = [x for x in xmls if x]
    if not xmls:
        return pd.DataFrame()

    frames = [parse_header_min(x) for x in xmls]
    h = pd.concat([f for f in frames if not f.empty], ignore_index=True) if frames else pd.DataFrame()

    if h.empty:
        return h

    rfc_u = rfc.upper()
    h = h[
        (h["uuid"].notna()) & (h["uuid"].astype(str) != "")
        & (h["tipo"].fillna("").str.upper() == "I")
        & (h["receptor_rfc"].fillna("").str.upper() == rfc_u)
        & (h["emisor_rfc"].fillna("").str.upper() != rfc_u)
    ].drop_duplicates(subset=["uuid"], keep="first")

    return h.reset_index(drop=True)


# =========================================================
# Métricas por intervalo (proveedores + participación)
# =========================================================

def metrics_by_interval(
    h: pd.DataFrame,
    interval_name: str,
    start_date: date,
    end_date: date,
    rfc_target: str,
    excluir_fx_desconocido: bool = True,
) -> Optional[pd.DataFrame]:
    if h is None or h.empty:
        return None

    hh = h.copy()

    # Ensure types consistent
    if "fecha" in hh.columns:
        hh["fecha"] = pd.to_datetime(hh["fecha"], errors="coerce").dt.date
    else:
        hh["fecha"] = pd.NaT

    rfc_u = rfc_target.upper()

    hh = hh[
        hh["fecha"].notna()
        & (hh["fecha"] >= start_date)
        & (hh["fecha"] <= end_date)
        & hh["emisor_rfc"].notna()
        & hh["receptor_rfc"].notna()
        & (hh["receptor_rfc"].astype(str).str.upper() == rfc_u)
        & (hh["emisor_rfc"].astype(str).str.upper() != rfc_u)
        & (hh["tipo"].fillna("").astype(str).str.upper() == "I")
    ].copy()

    if hh.empty:
        return None

    hh["moneda"] = hh.get("moneda", pd.Series([None]*len(hh))).astype(str).str.upper()
    hh["metodo"] = hh.get("metodo", pd.Series([None]*len(hh))).astype(str).str.upper()
    hh["tipo_cambio"] = pd.to_numeric(hh.get("tipo_cambio"), errors="coerce")
    hh["total"] = pd.to_numeric(hh.get("total"), errors="coerce")

    # monto_mxn logic matches case_when in R
    monto_mxn = []
    for mon, fx, tot in zip(hh["moneda"], hh["tipo_cambio"], hh["total"]):
        if pd.isna(tot):
            monto_mxn.append(None)
            continue
        if (mon is None) or (str(mon).upper() == "MXN") or (str(mon).strip() == "NAN"):
            monto_mxn.append(float(tot))
        elif pd.notna(fx) and float(fx) > 0 and pd.isfinite(fx):
            monto_mxn.append(float(tot) * float(fx))
        else:
            monto_mxn.append(None)
    hh["monto_mxn"] = pd.to_numeric(pd.Series(monto_mxn), errors="coerce")

    if excluir_fx_desconocido:
        hh = hh[hh["monto_mxn"].notna()].copy()

    if hh.empty:
        return None

    # columns
    monto_col = f"Monto total facturas ({interval_name})"
    part_col = f"Participación ({interval_name})"
    ct_total = f"Conteo total facturas ({interval_name})"
    ct_ppd = f"Conteo facturas PPD ({interval_name})"
    monto_ppd = f"Monto facturado PPD ({interval_name})"
    ct_pue = f"Conteo facturas PUE ({interval_name})"
    monto_pue = f"Monto facturado PUE ({interval_name})"

    hh["emisor_nombre"] = hh["emisor_nombre"].where(hh["emisor_nombre"].notna(), hh["emisor_rfc"])

    grp = hh.groupby(["emisor_rfc", "emisor_nombre"], dropna=False)

    agg = grp.agg(
        **{
            ct_total: ("uuid", "count"),
            monto_col: ("monto_mxn", "sum"),
            ct_ppd: ("metodo", lambda s: int((s == "PPD").sum())),
            monto_ppd: ("monto_mxn", lambda s: float(s[hh.loc[s.index, "metodo"] == "PPD"].sum())),
            ct_pue: ("metodo", lambda s: int((s == "PUE").sum())),
            monto_pue: ("monto_mxn", lambda s: float(s[hh.loc[s.index, "metodo"] == "PUE"].sum())),
        }
    ).reset_index()

    total_sum = float(agg[monto_col].sum()) if monto_col in agg.columns else 0.0
    agg[part_col] = (agg[monto_col] / total_sum) if total_sum > 0 else 0.0

    return agg


# =========================================================
# Excel: tabla con encabezados agrupados por intervalo
# =========================================================

def write_grouped_table(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    intervals: List[str],
    start_row: int = 5,
) -> None:
    """
    Writes a table with:
    - a group header row (merged cells) for each interval
    - a second header row with actual column names
    - data rows
    - basic number formats similar to R
    """
    ws = wb[sheet_name]

    if df is None or df.empty:
        ws.cell(row=start_row, column=1, value="Sin datos")
        return

    # Ensure provider columns in expected names
    fixed_cols = []
    for c in ["Proveedor_RFC", "Proveedor_Nombre", "Participación (%)"]:
        if c in df.columns:
            fixed_cols.append(c)

    # Build per-interval column groups based on exact naming patterns produced by metrics_by_interval joins
    def cols_for_interval(lbl: str) -> List[str]:
        wanted = [
            f"Conteo total facturas ({lbl})",
            f"Monto total facturas ({lbl})",
            f"Conteo facturas PPD ({lbl})",
            f"Monto facturado PPD ({lbl})",
            f"Conteo facturas PUE ({lbl})",
            f"Monto facturado PUE ({lbl})",
            f"Participación ({lbl})",
        ]
        return [c for c in wanted if c in df.columns]

    interval_cols: List[Tuple[str, List[str]]] = [(lbl, cols_for_interval(lbl)) for lbl in intervals]
    interval_cols = [(lbl, cols) for lbl, cols in interval_cols if cols]

    ordered_cols: List[str] = fixed_cols[:]
    for _, cols in interval_cols:
        ordered_cols.extend(cols)

    # Add any remaining columns (shouldn't happen, but keep safe)
    remaining = [c for c in df.columns if c not in ordered_cols]
    ordered_cols.extend(remaining)

    df2 = df[ordered_cols].copy()

    # Styles
    fill_group = PatternFill("solid", fgColor="0B2E4E")  # navy
    fill_head = PatternFill("solid", fgColor="E5E7EB")   # light border-ish
    font_group = Font(color="FFFFFF", bold=True)
    font_head = Font(bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row indices
    r_group = start_row
    r_head = start_row + 1
    r_data = start_row + 2

    # Write group header row (merged cells)
    col_idx = 1
    # Fixed group header: "Proveedor"
    if fixed_cols:
        span = len(fixed_cols)
        ws.merge_cells(start_row=r_group, start_column=col_idx, end_row=r_group, end_column=col_idx + span - 1)
        cell = ws.cell(row=r_group, column=col_idx, value="Proveedor")
        cell.fill = fill_group
        cell.font = font_group
        cell.alignment = align_center
        cell.border = border
        for c in range(col_idx, col_idx + span):
            ws.cell(row=r_group, column=c).border = border
            ws.cell(row=r_group, column=c).fill = fill_group
        col_idx += span

    # Interval groups
    for lbl, cols in interval_cols:
        if not cols:
            continue
        span = len(cols)
        ws.merge_cells(start_row=r_group, start_column=col_idx, end_row=r_group, end_column=col_idx + span - 1)
        cell = ws.cell(row=r_group, column=col_idx, value=lbl)
        cell.fill = fill_group
        cell.font = font_group
        cell.alignment = align_center
        cell.border = border
        for c in range(col_idx, col_idx + span):
            ws.cell(row=r_group, column=c).border = border
            ws.cell(row=r_group, column=c).fill = fill_group
        col_idx += span

    # Header row (column names)
    for j, col in enumerate(df2.columns, start=1):
        cell = ws.cell(row=r_head, column=j, value=col)
        cell.fill = fill_head
        cell.font = font_head
        cell.alignment = align_center
        cell.border = border

    # Data rows
    for i, (_, row) in enumerate(df2.iterrows(), start=0):
        for j, col in enumerate(df2.columns, start=1):
            val = row[col]
            cell = ws.cell(row=r_data + i, column=j, value=None if pd.isna(val) else val)
            cell.border = border
            # formatting
            if isinstance(val, (int, float)) and ("Participación" in col):
                cell.number_format = "0.00%"
            elif isinstance(val, (int, float)) and ("Monto" in col):
                cell.number_format = u'"$"#,##0.00'
            elif isinstance(val, (int, float)) and ("Conteo" in col):
                cell.number_format = "0"
            cell.alignment = Alignment(vertical="top")

    # Freeze panes at header
    ws.freeze_panes = ws.cell(row=r_data, column=1)

    # Auto width (basic)
    for j, col in enumerate(df2.columns, start=1):
        max_len = max([len(str(col))] + [len(str(x)) for x in df2[col].head(200).astype(str).tolist()])
        ws.column_dimensions[get_column_letter(j)].width = min(55, max(10, max_len + 2))


def build_excel_for_rfcs(
    base_url: str,
    api_key: str,
    rfcs: List[str],
    source: str,
    intervals: List[str],
    excluir_fx: bool,
) -> bytes:
    """
    Mirrors the Shiny downloadHandler: 1 sheet per RFC, info header, grouped table at row 5.
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    max_days = max(INTERVAL_DEFS[i] for i in intervals) if intervals else 365

    for rfc in rfcs:
        dto = today_utc()
        dfrom = dto - timedelta(days=max_days)

        if source == "api":
            h = list_invoices_headers_api(base_url, api_key, rfc, dfrom, dto)
        else:
            h = headers_from_xml(base_url, api_key, rfc, dfrom, dto)

        sheet_name = rfc[:31]
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        # Header info (rows 1-3)
        ws.cell(row=1, column=1, value=f"RFC: {rfc}")
        ws.cell(row=2, column=1, value=f"Fuente: {'API (facturas)' if source == 'api' else 'XML'}")
        ws.cell(row=3, column=1, value=f"Intervalos: {', '.join(intervals)}")

        if h is None or h.empty:
            ws.cell(row=5, column=1, value="Sin datos para los intervalos seleccionados.")
            continue

        blocks: List[pd.DataFrame] = []
        for lbl in intervals:
            days_back = INTERVAL_DEFS[lbl]
            start = dto - timedelta(days=days_back)
            b = metrics_by_interval(h, lbl, start, dto, rfc, excluir_fx_desconocido=excluir_fx)
            if b is not None and not b.empty:
                blocks.append(b)

        if not blocks:
            ws.cell(row=5, column=1, value="Sin datos para los intervalos seleccionados.")
            continue

        # Full join by emisor_rfc + emisor_nombre
        out = blocks[0]
        for k in range(1, len(blocks)):
            out = out.merge(blocks[k], on=["emisor_rfc", "emisor_nombre"], how="outer")

        # Columna fija Participación (%) del intervalo de referencia = intervals[0]
        ref_lbl = intervals[0]
        ref_part_col = f"Participación ({ref_lbl})"
        ref_total_col = f"Monto total facturas ({ref_lbl})"
        if ref_part_col in out.columns:
            out["Participación (%)"] = out[ref_part_col]
        elif ref_total_col in out.columns:
            ts = float(pd.to_numeric(out[ref_total_col], errors="coerce").sum())
            out["Participación (%)"] = (out[ref_total_col] / ts) if ts > 0 else 0.0
        else:
            out["Participación (%)"] = 0.0

        # Rename provider columns and relocate Participación (%) after name
        out = out.rename(columns={"emisor_rfc": "Proveedor_RFC", "emisor_nombre": "Proveedor_Nombre"})
        # relocate
        cols = list(out.columns)
        if "Participación (%)" in cols:
            cols.remove("Participación (%)")
            # place after Proveedor_Nombre if present
            if "Proveedor_Nombre" in cols:
                idx = cols.index("Proveedor_Nombre") + 1
            else:
                idx = 1
            cols.insert(idx, "Participación (%)")
            out = out[cols]

        write_grouped_table(wb, sheet_name, out, intervals, start_row=5)

    # Serialize to bytes
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
